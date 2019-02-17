
"""
POLARIS - Poloniex Cryptocurrency Directional Trading Bot

Created on Fri Aug 11 07:38:39 2017
Modified on Fri Sep 22 2017
Modified on Thr Oct 5 2017


@author: Pfmoro - Based on code by bwentzloff
"""


import time
import sys, getopt
import datetime
import numpy
import urllib
import urllib.request
import openpyxl
import json
import hmac,hashlib
import calendar
from sklearn.linear_model import LinearRegression


def createTimeStamp(datestr, format="%Y-%m-%d %H:%M:%S"):
    return time.mktime(time.strptime(datestr, format))

class poloniex:
    def __init__(self, APIKey, Secret):
        self.APIKey = APIKey
        self.Secret = Secret

    def post_process(self, before):
        after = before

        # Add timestamps if there isnt one but is a datetime
        if('return' in after):
            if(isinstance(after['return'], list)):
                for x in xrange(0, len(after['return'])):
                    if(isinstance(after['return'][x], dict)):
                        if('datetime' in after['return'][x] and 'timestamp' not in after['return'][x]):
                            after['return'][x]['timestamp'] = float(createTimeStamp(after['return'][x]['datetime']))
                            
        return after

    def api_query(self, command, req={}):

        if(command == "returnTicker" or command == "return24Volume"):
            ret = urllib.request.urlopen(urllib.request.Request(('https://poloniex.com/public?command=' + command)))
            return json.loads(ret.read())
        elif(command == "returnOrderBook"):
            ret = urllib.request.urlopen(urllib.request.Request(('https://poloniex.com/public?command=' + command + '&currencyPair=' + str(req['currencyPair']))))
            return json.loads(ret.read())
        elif(command == "returnMarketTradeHistory"):
            ret = urllib.request.urlopen(urllib.request.Request(('https://poloniex.com/public?command=' + "returnTradeHistory" + '&currencyPair=' + str(req['currencyPair']))))
            return json.loads(ret.read())
        elif(command == "returnChartData"):
            ret = urllib.request.urlopen(urllib.request.Request(('https://poloniex.com/public?command=returnChartData&currencyPair=' + str(req['currencyPair']) + '&start=' + str(req['start']) + '&end=' + str(req['end']) + '&period=' + str(req['period']))))
            return json.loads(ret.read())
        else:
            req['command'] = command
            req['nonce'] = int(time.time()*1000)
            post_data = urllib.parse.urlencode(req)
            
            
            sign = hmac.new(self.Secret, post_data.encode('utf-8'), hashlib.sha512).hexdigest()
            headers = {
                'Sign': sign,
                'Key': self.APIKey
            }

            ret = urllib.request.urlopen(urllib.request.Request('https://poloniex.com/tradingApi', post_data.encode('utf-8'), headers))
            jsonRet = json.loads(ret.read())
            return self.post_process(jsonRet)


    def returnTicker(self):
        return self.api_query("returnTicker")

    def return24Volume(self):
        return self.api_query("return24Volume")

    def returnOrderBook (self, currencyPair):
        return self.api_query("returnOrderBook", {'currencyPair': currencyPair})

    def returnMarketTradeHistory (self, currencyPair):
        return self.api_query("returnMarketTradeHistory", {'currencyPair': currencyPair})


    # Returns all of your balances.
    # Outputs: 
    # {"BTC":"0.59098578","LTC":"3.31117268", ... }
    def returnBalances(self):
        return self.api_query('returnBalances')

    # Returns your open orders for a given market, specified by the "currencyPair" POST parameter, e.g. "BTC_XCP"
    # Inputs:
    # currencyPair  The currency pair e.g. "BTC_XCP"
    # Outputs: 
    # orderNumber   The order number
    # type          sell or buy
    # rate          Price the order is selling or buying at
    # Amount        Quantity of order
    # total         Total value of order (price * quantity)
    def returnOpenOrders(self,currencyPair):
        return self.api_query('returnOpenOrders',{"currencyPair":currencyPair})


    # Returns your trade history for a given market, specified by the "currencyPair" POST parameter
    # Inputs:
    # currencyPair  The currency pair e.g. "BTC_XCP"
    # Outputs: 
    # date          Date in the form: "2014-02-19 03:44:59"
    # rate          Price the order is selling or buying at
    # amount        Quantity of order
    # total         Total value of order (price * quantity)
    # type          sell or buy
    def returnTradeHistory(self,currencyPair):
        return self.api_query('returnTradeHistory',{"currencyPair":currencyPair})

    # Places a buy order in a given market. Required POST parameters are "currencyPair", "rate", and "amount". If successful, the method will return the order number.
    # Inputs:
    # currencyPair  The curreny pair
    # rate          price the order is buying at
    # amount        Amount of coins to buy
    # Outputs: 
    # orderNumber   The order number
    def buy(self,currencyPair,rate,amount):
        return self.api_query('buy',{"currencyPair":currencyPair,"rate":rate,"amount":amount})

    # Places a sell order in a given market. Required POST parameters are "currencyPair", "rate", and "amount". If successful, the method will return the order number.
    # Inputs:
    # currencyPair  The curreny pair
    # rate          price the order is selling at
    # amount        Amount of coins to sell
    # Outputs: 
    # orderNumber   The order number
    def sell(self,currencyPair,rate,amount):
        return self.api_query('sell',{"currencyPair":currencyPair,"rate":rate,"amount":amount})

    # Cancels an order you have placed in a given market. Required POST parameters are "currencyPair" and "orderNumber".
    # Inputs:
    # currencyPair  The curreny pair
    # orderNumber   The order number to cancel
    # Outputs: 
    # succes        1 or 0
    def cancel(self,currencyPair,orderNumber):
        return self.api_query('cancelOrder',{"currencyPair":currencyPair,"orderNumber":orderNumber})

    # Immediately places a withdrawal for a given currency, with no email confirmation. In order to use this method, the withdrawal privilege must be enabled for your API key. Required POST parameters are "currency", "amount", and "address". Sample output: {"response":"Withdrew 2398 NXT."} 
    # Inputs:
    # currency      The currency to withdraw
    # amount        The amount of this coin to withdraw
    # address       The withdrawal address
    # Outputs: 
    # response      Text containing message about the withdrawal
    def withdraw(self, currency, amount, address):
        return self.api_query('withdraw',{"currency":currency, "amount":amount, "address":address})

def main(argv):
	#Parâmetros de inicialização:
	TradeAmount = 0.001
	period = 1800 # 300,900,1800,7200,14400, or 86400 seconds
	pair = "BTC_ETH" #"BTC_LTC" "BTC_XMR" "BTC_ETH" "BTC_ETC" "USDT_BTC" "BTC_DASH" "BTC_STEEM", "USDT_XRP" et all
	TradeLock = "Long" #Valores possíveis:"Long" ou "Short”
   #/*Se ajustada para “Long”,o Usuário entende que a tendência de Longo prazo e alta e isso Proibe o robô de operar Short. Ae o usuário entender que deve operar Short, trocar para “Short”, essa avaliação deve ocorrer UMA VEZ POR SEMANA. */
	ChangeTradeLock = False #se true, Polaris tem autonomia para mudar o Trade Lock, se false, polaris somente fará trades se o o seu trade lock bater com o do usuário, se outro valor a avaiação do polaris é desconsiderada
	TrendTestLimit = 0 #tangente do angulo que sera usado para analise de tendencia
	lengthOfMA = 13 #de 13 a 26 periodos
	longTermengthOfMA = 26 #de 13 a 26 periodos    
	LongTermAnalysisPeriod  = 26 #recomendado 96
	#TrendShortMAPeriod = 26#periodos para medias moveis de analise de longo prazo, curta menor que longa e ambas menores que LongTermAnalysisPeriod
	#TrendLongMAPeriod = 52
	LongPeriod =14400 # Periodo de analise da tendencia de longo prazo, deve respeitar as restricoes da poloniex  
	AllowLossOnStop = True #se false, só ativa o stop loss quando não for possível haver prejuizo. pode deixar o polaris travado em um trade 
    
    #Parametros facultativos
	tradePlaced = False
	typeOfTrade = ""    
	StopLoss = 0
	StopGain = 0
	EnterPrice = 0
    #/*preço de entrada no Trade, para medição de performance do Polaris*/ 
    
    #definições padrões:
	UserTradeLock = ""
	PolarisTradeLock =""
	X_Norm =[]
	prices = []
	LongTermPrices =[]
	TrendAnalysisPrices =[]
	TrendAnalysisPricesNorm =[]
	X_Reg =[]
	currentMovingAverage = 0;
	pendingCount = 0
	OneSigma = 0;    	
	startTime = False
	endTime = False
	MovingAverages =[]
	historicalData = False

	dataDate = ""

	Oscilator = 0.5
	orderNumber = ""
	lastPairPrice = 0
	dataPoints = []
	TradePendingExec = False
	PersonalBook =""
	localMax = []
	currentResistance = 0.018

	Result = 0
	LossCount = 0
	DataBase = 'YOUR FULL FILE PATH FOR AN SPREADSHEET'
	TradeFee = 0.0025 #Poloniex uses 0.25% and 0.15% as trade fees
    
	try:
		opts, args = getopt.getopt(argv,"hp:c:n:s:e:",["period=","currency=","points="])
	except getopt.GetoptError:
		print ("Polaris.py -p <period length> -c <currency pair> -n <period of moving average>")
		sys.exit(2)

	for opt, arg in opts:
		if opt == '-h':
			print ('Polaris.py -p <period length> -c <currency pair> -n <period of moving average>')
			sys.exit()
		elif opt in ("-p", "--period"):
			if (int(arg) in [300,900,1800,7200,14400,86400]):
				period = arg
			else:
				print ('Poloniex requires periods in 300,900,1800,7200,14400, or 86400 second increments')
				sys.exit(2)
		elif opt in ("-c", "--currency"):
			pair = arg
		elif opt in ("-n", "--points"):
			lengthOfMA = int(arg)
		elif opt in ("-s"):
			startTime = arg
		elif opt in ("-e"):
			endTime = arg

	Key = ('YOUR KEY').encode('utf-8')
	Secret = ('YOUR SECRET').encode('utf-8')
	conn = poloniex(Key,Secret)
	print ("Polaris Online!")
    
	UserTradeLock = TradeLock
	PolarisTradeLock =TradeLock
	print("Recovering Historical Data...")
	flag =True
	cont = 0
	while flag ==True:
		cont = cont+1
		HendTime = str(calendar.timegm(datetime.datetime.utcnow().utctimetuple())-period*(cont-1))
		HstartTime=str(calendar.timegm(datetime.datetime.utcnow().utctimetuple())-period*cont)
        
		HistoricalVector= conn.api_query("returnChartData",{"currencyPair":pair,"start":HstartTime,"end":HendTime,"period":period})
		if len(prices)< lengthOfMA: 
		 prices.append(float(str(str(HistoricalVector)[(str(HistoricalVector).find('weightedAverage')+3  + len('weightedAverage')):str(HistoricalVector).find('}')] )))
		if len(LongTermPrices)< 5*lengthOfMA:
		 LongTermPrices.append(float(str(str(HistoricalVector)[(str(HistoricalVector).find('weightedAverage')+3  + len('weightedAverage')):str(HistoricalVector).find('}')] )))
		if len(LongTermPrices)>= 5*lengthOfMA and len(prices)>= lengthOfMA:
		 flag = False
		print("Prices array size: " + str(len(prices)) + " LongTermPrices array size: "+ str(len(LongTermPrices)) )    
		if (len(prices) > 0):
						if float(len(prices)) < lengthOfMA:
						 currentMovingAverage = numpy.average(prices)
						else:
						 currentMovingAverage = (float(prices[len(prices)-1])  - float(currentMovingAverage)) * (2/(float(lengthOfMA) +1)) + float(currentMovingAverage)
		else:
						 currentMovingAverage = float(prices[len(prices)-1])
	print("Historical Data Recovery is complete")  
	flag =True
	cont = 0
	print("Recovering Long Term Data...")    
	while flag ==True:   
		cont = cont+1
		HendTime = str(calendar.timegm(datetime.datetime.utcnow().utctimetuple())-LongPeriod*(cont-1))
		HstartTime=str(calendar.timegm(datetime.datetime.utcnow().utctimetuple())-LongPeriod*cont)        
		PriceInitVector = conn.api_query("returnChartData",{"currencyPair":pair,"start":HstartTime,"end":HendTime,"period":LongPeriod})
		X_Reg.append(cont)
		TrendAnalysisPrices.append(float(str(str(PriceInitVector)[(str(PriceInitVector).find('weightedAverage')+3  + len('weightedAverage')):str(PriceInitVector).find('}')] )))
        
		if (len(TrendAnalysisPrices) > 0):
						if float(len(TrendAnalysisPrices)) < longTermengthOfMA:
						 LTMovingAverage = numpy.average(TrendAnalysisPrices)
						else:
						 LTMovingAverage = (float(TrendAnalysisPrices[len(TrendAnalysisPrices)-1])  - float(LTMovingAverage)) * (2/(float(longTermengthOfMA) +1)) + float(LTMovingAverage)
		else:
						 LTMovingAverage = TrendAnalysisPrices[len(TrendAnalysisPrices)-1]
		MovingAverages.append(LTMovingAverage)

		if len(TrendAnalysisPrices)   >= LongTermAnalysisPeriod: flag = False
		print("Trend analysis Array size: " + str(len(TrendAnalysisPrices))) 
	MovingAverages.reverse()    
	print("Long Term Data Recovery is complete")     
	HendTime = str(calendar.timegm(datetime.datetime.utcnow().utctimetuple()))       
	cont = 0
	while cont<len(X_Reg): 
	  X_Norm.append((X_Reg[cont]- numpy.average(X_Reg))/numpy.std(X_Reg))
	  cont=cont+1
	#  TrendAnalysisPricesNorm.append((TrendAnalysisPrices[cont]- numpy.average(TrendAnalysisPrices))/numpy.std(TrendAnalysisPrices))
        
	output = open("output.html",'w')
	output.truncate()
#	output.write("""<html><head><script type="text/javascript" src="https://www.gstatic.com/charts/loader.js"></script><script type="text/javascript">google.charts.load('current', {'packages':['corechart']});google.charts.setOnLoadCallback(drawChart);function drawChart() {var data = new google.visualization.DataTable();data.addColumn('string', 'time');data.addColumn('number', 'value');data.addColumn({type: 'string', role:'annotation'});data.addColumn({type: 'string', role:'annotationText'});data.addColumn('number', 'trend');data.addRows([""")
   
	try:      
			 PersonalBook = conn.returnOpenOrders(pair)
			 if PersonalBook ==[]:
			  TradePendingExec = False
			  print("No trades pending execution")
			 else:
			  TradePendingExec = True
			  print("Warning: There are trades pending execution")
	except:
			 print("Unable to check Personal book") 

	if (startTime):
		historicalData = conn.api_query("returnChartData",{"currencyPair":pair,"start":startTime,"end":endTime,"period":period})

	while True:
		if (startTime and historicalData):
			nextDataPoint = historicalData.pop(0)
			lastPairPrice = nextDataPoint['weightedAverage']
			dataDate = datetime.datetime.fromtimestamp(int(nextDataPoint['date'])).strftime('%Y-%m-%d %H:%M:%S')
		elif(startTime and not historicalData):
			for point in dataPoints:
				output.write("['"+point['date']+"',"+point['price']+","+point['label']+","+point['desc']+","+point['trend'])
				output.write("],\n")
			output.write("""]);var options = {title: 'Price Chart',legend: { position: 'bottom' }};var chart = new google.visualization.LineChart(document.getElementById('curve_chart'));chart.draw(data, options);}</script></head><body><div id="curve_chart" style="width: 100%; height: 100%"></div></body></html>""")
			exit()
		else:
			try:
			 currentValues = conn.api_query("returnTicker")
			 lastPairPrice = currentValues[pair]["last"]
			 dataDate = datetime.datetime.now()
			except:
			 print("Unable to connect to Poloniex, please check your Internet")  
            
		dataPoints.append({'date':dataDate, 'price': str(lastPairPrice), 'trend': str(currentResistance), 'label': 'null', 'desc': 'null'})
       
	#	Wb = openpyxl.load_workbook('CryptoPoloData.xlsx') 
	#	sheet=Wb.get_sheet_by_name(pair)
	#	for row in range(2,sheet.max_row + 1):
	#	 LongTermPrices.append(sheet['B'+ str(row)])
	#	for row in range(sheet.max_row - lengthOfMA, sheet.max_row + 1):       
	#	 prices.append(sheet['B'+ str(row)])
	#	Wb.close
		cont = 0
		RupturaBaixa = 0
		RupturaAlta = 0        
		while cont < len(LongTermPrices)-1:
		 #print('cont: ' + str(cont))
		 if (LongTermPrices[cont+1] - LongTermPrices[cont])< RupturaBaixa: RupturaBaixa = (LongTermPrices[cont] - LongTermPrices[cont+1])     
		 if (LongTermPrices[cont+1] - LongTermPrices[cont])> RupturaAlta: RupturaAlta = (LongTermPrices[cont] - LongTermPrices[cont+1])     
		 if TradeLock == "Long": OneSigma = RupturaBaixa           
		 if TradeLock == "Short": OneSigma = RupturaAlta 
		 cont= cont+1         
            
		Volatility = float(numpy.std(LongTermPrices)) #Cálculo do desvio padrão para utilizar juntamente com a MME
		print("Current Volatility is: " + str(Volatility))
       
		if ( (len(dataPoints) > 2) and (dataPoints[-2]['price'] > dataPoints[-1]['price']) and (dataPoints[-2]['price'] > dataPoints[-3]['price']) ):
			dataPoints[-2]['label'] = "'MAX'"
			dataPoints[-2]['desc'] = "'This is a local maximum'"
			
			numberOfSimilarLocalMaxes = 0
			for oldMax in localMax:
				if ( (float(oldMax) > (float(dataPoints[-2]['price']) - .0001) ) and (float(oldMax) < (float(dataPoints[-2]['price']) + .0001) ) ):
					numberOfSimilarLocalMaxes = numberOfSimilarLocalMaxes + 1

			if (numberOfSimilarLocalMaxes > 2):
				currentResistance = dataPoints[-2]['price']
				dataPoints[-2]['trend'] = dataPoints[-2]['price']
				dataPoints[-1]['trend'] = dataPoints[-2]['price']

			localMax.append(dataPoints[-2]['price'])


                
	
		if (len(prices) > 1):    
			 previousPrice = float(prices[-1])
		else:
			 previousPrice =  lastPairPrice     
             
		if (len(prices) > 0):
						if float(len(prices)) < lengthOfMA:
						 currentMovingAverage = numpy.average(prices)
						else:
						 currentMovingAverage = (float(lastPairPrice)  - float(currentMovingAverage)) * (2/(float(lengthOfMA) +1)) + float(currentMovingAverage)
		else:
						 currentMovingAverage = float(lastPairPrice)                          
             
		if (len(prices) > 0):
						 minimus=min(float(lastPairPrice),float(min(prices)))
						 maximus=max(float(lastPairPrice),float(max(prices)))
						 Oscilator = ((float(lastPairPrice) - minimus)/(maximus - minimus))
		else:
						 Oscilator = 0.5
       
                
		if tradePlaced == True: 
		 if typeOfTrade =="Long" and (float(lastPairPrice) > float(StopLoss) + 2*float(OneSigma)):
		  print("Stop Loss Updated")    
		  StopLoss = float(StopLoss) + float(OneSigma)  
		 if typeOfTrade =="Short" and float(lastPairPrice) < float(StopLoss) - 2*float(OneSigma):
		  print("Stop Loss Updated")    
		  StopLoss = float(StopLoss) - float(OneSigma)          
		 print("Active trade - " " Target: "+ str(StopGain) + " Stop Loss: " + str(StopLoss))
                         
                         
                         
		if ((not tradePlaced) and (TradePendingExec == False) and 	LossCount <3):
				if (( (float(lastPairPrice) >= float(currentMovingAverage)) or float(Oscilator) > 0.8) and (TradeLock == "Short") and float(previousPrice) > float(lastPairPrice) ):
					print("SELL ORDER")
					orderNumber = conn.sell(pair,lastPairPrice,TradeAmount)
					#print("Order Number: "+ orderNumber)
					EnterPrice = lastPairPrice
					tradePlaced = True
					typeOfTrade = "Short"
					StopLoss = float(lastPairPrice) + float(OneSigma)
					StopGain = float(lastPairPrice) - 3*float(OneSigma)
					print("Target is: " +str(StopGain) + " StopLoss is: " + str(StopLoss) )
				elif (( (float(lastPairPrice) <= float(currentMovingAverage)) or float(Oscilator) < 0.2) and (TradeLock == "Long") and float(previousPrice) < float(lastPairPrice) ):
					print("BUY ORDER")
					orderNumber = conn.buy(pair,lastPairPrice,TradeAmount)
					#print("Order Number: "+ orderNumber)                    
					EnterPrice = lastPairPrice
					tradePlaced = True
					typeOfTrade = "Long"
					StopLoss = float(lastPairPrice) - float(OneSigma)                  
					StopGain = float(lastPairPrice) + 3*float(OneSigma)
					print("Target is: " +str(StopGain) + " StopLoss is: " + str(StopLoss) )
		elif (typeOfTrade == "Short"):
				if float(lastPairPrice) < (StopLoss - 1.5*float(OneSigma)): StopLoss = lastPairPrice   
				if ( float(lastPairPrice) < StopGain):
					print("EXIT TRADE - Stop Gain")
					try:
					 try:#Cancela o trade, se ele não foi executado
					  conn.cancel(pair,orderNumber)
					#print("Order Number: "+ orderNumber)     retorna:  TypeError: must be str, not dict                
					 except:#inverte a posição)
					  orderNumber = conn.buy(pair,lastPairPrice,float(TradeAmount)*(1-float(TradeFee)))
					 Result = (float(lastPairPrice) - float(EnterPrice))*TradeAmount
					 print("Result: "+ str(Result ))
					 LossCount = 0
					 tradePlaced = False
					 typeOfTrade = False
					except:  
					 print("Unable to unwind trade")     
				if ( AllowLossOnStop == True and (float(lastPairPrice) > float(StopLoss)) or (AllowLossOnStop == False and (float(lastPairPrice) > float(StopLoss)) and (float(lastPairPrice) < float(EnterPrice)*(1-float(TradeFee))))):
					print("EXIT TRADE - Stop Loss")
					try:
					 try:#Cancela o trade, se ele não foi executado
					  conn.cancel(pair,orderNumber)
					#print("Order Number: "+ orderNumber)     retorna:  TypeError: must be str, not dict                
					 except:#inverte a posição)
					  orderNumber = conn.buy(pair,lastPairPrice,float(TradeAmount)*(1-float(TradeFee)))
					 Result =(float(lastPairPrice) - float(EnterPrice))*TradeAmount
					 LossCount = LossCount +1
					 print("Result: "+ str(Result ))
					 if LossCount >= 3: print("Trading is Halted until next trend evaluation due to three sucessive losses")
					 tradePlaced = False
					 typeOfTrade = False
					except:  
					 print("Unable to unwind trade")                          
		elif (typeOfTrade == "Long"):
				if float(lastPairPrice) > (StopLoss + 1.5*float(OneSigma)): StopLoss = lastPairPrice           
				if (float(lastPairPrice) > StopGain):
					print("EXIT TRADE - Stop Gain")    
					try:
					 try:#Cancela o trade, se ele não foi executado
					  conn.cancel(pair,orderNumber)
					#print("Order Number: "+ orderNumber)   retorna:  TypeError: must be str, not dict                   
					 except: #inverte a posição
					  orderNumber = conn.sell(pair,lastPairPrice,float(TradeAmount)*(1-float(TradeFee)))                  
					 Result = (float(lastPairPrice) - float(EnterPrice))*TradeAmount
					 LossCount = 0        
					 print("Result: "+ str(Result) )
					 tradePlaced = False
					 typeOfTrade = False
					except:  
					 print("Unable to unwind trade")                     
				if ( AllowLossOnStop == True and float(lastPairPrice) < float(StopLoss ) or (AllowLossOnStop == False and float(lastPairPrice) < float(StopLoss) and (float(lastPairPrice) > float(EnterPrice)*(1+float(TradeFee))))):
					print("EXIT TRADE - Stop Loss")    
					try:
					 try:#Cancela o trade, se ele não foi executado
					  conn.cancel(pair,orderNumber)
					#print("Order Number: "+ orderNumber)   retorna:  TypeError: must be str, not dict                   
					 except: #inverte a posição
					  orderNumber = conn.sell(pair,lastPairPrice,float(TradeAmount)*(1-float(TradeFee)))                     
					 Result = (float(lastPairPrice) - float(EnterPrice))*TradeAmount
					 if Result < 0: LossCount = LossCount +1
					 print("Result: "+ str(Result ))
					 if LossCount >= 3: print("Trading is Halted until next trend evaluation due to three sucessive losses")
					 tradePlaced = False
					 typeOfTrade = False
					except:  
					 print("Unable to unwind trade")     
                     
                     
                     
		else:
			previousPrice = 0
		try:      
			 PersonalBook = conn.returnOpenOrders(pair)
			 if PersonalBook ==[]:
			  TradePendingExec = False
			  pendingCount = 0 
			  print("No trades pending execution")
			 else:
			  TradePendingExec = True
			  pendingCount = pendingCount +1
			  print("Warning: There are trades pending execution")
			  if pendingCount> 2 and tradePlaced == True:
			   tradePlaced = True
			   try:
			    conn.cancel(pair,orderNumber)
			    pendingCount = 0
			    print("Unable to unwind trade due to current market conditions: operation is reactivated")   
			   except:                
			    print("Unable to reactivate operation")             
		except:
			 print("Unable to check Personal book") 
         
		print( str(dataDate)+ " Period: " + str(period) +" "+ pair +": " + str(lastPairPrice) + " Moving Average: " + str(currentMovingAverage) + " Oscilator: "+ str(Oscilator))
               
        # Guardar Historico em arquivo excel.
		if Result != 0:        
		 Wb = openpyxl.load_workbook(DataBase)
		 sheet = Wb.get_sheet_by_name(pair)
		 sheet['A'+ str(sheet.max_row + 1)] = str(dataDate)
		 sheet['B'+ str(sheet.max_row)] = float(lastPairPrice)
		 sheet['C'+ str(sheet.max_row)] = float(EnterPrice)
		 sheet['D'+ str(sheet.max_row)] = float(Result)
		 Result = 0
		 Wb.save(DataBase)
		 Wb.close
 
		if (calendar.timegm(datetime.datetime.utcnow().utctimetuple()) - int(HendTime)) >= LongPeriod:
			HendTime = str(calendar.timegm(datetime.datetime.utcnow().utctimetuple()))
			HstartTime=str(calendar.timegm(datetime.datetime.utcnow().utctimetuple())-LongPeriod)
			PriceInitVector = conn.api_query("returnChartData",{"currencyPair":pair,"start":HstartTime,"end":HendTime,"period":LongPeriod})
			TrendAnalysisPrices.append(float(str(str(PriceInitVector)[(str(PriceInitVector).find('weightedAverage')+3  + len('weightedAverage')):str(PriceInitVector).find('}')] )))
			if (len(TrendAnalysisPrices) > 0):
						if float(len(TrendAnalysisPrices)) < longTermengthOfMA:
						 LTMovingAverage = numpy.average(TrendAnalysisPrices)
						else:
						 LTMovingAverage = (float(TrendAnalysisPrices[-1])  - float(LTMovingAverage)) * (2/(float(longTermengthOfMA) +1)) + float(LTMovingAverage)  
			else:
						 LTMovingAverage = TrendAnalysisPrices[-1]
			MovingAverages.append(LTMovingAverage)
			if LossCount >= 3: 
			 LossCount = 0
			 print("Trading operations were resumed")   
			print("Long Term trend array updated")
			#cont=0
			#while cont<len(X_Reg): 
			#  X_Norm.append((X_Reg[cont]- numpy.average(X_Reg))/numpy.std(X_Reg))
			  #TrendAnalysisPricesNorm.append((TrendAnalysisPrices[cont]- numpy.average(TrendAnalysisPrices))/numpy.std(TrendAnalysisPrices))
         
		     
		LongTermPrices.append(float(lastPairPrice))
		prices.append(float(lastPairPrice))
		prices = prices[-lengthOfMA:]
		LongTermPrices =LongTermPrices[-5*lengthOfMA:]
		MovingAverages = MovingAverages[-longTermengthOfMA:]
		TrendAnalysisPrices =TrendAnalysisPrices[-LongTermAnalysisPeriod:]
		TrendAnalysisPricesNorm =TrendAnalysisPricesNorm[-LongTermAnalysisPeriod:]        
	#	X_Reg =X_Reg[-LongTermAnalysisPeriod:]
	#	X_Norm =X_Norm[-LongTermAnalysisPeriod:]        
    
		y1 =[]
		y2=[]
# Análise de Tendencia de longo prazo
# 1 - Regressao Linear do vetor de preços
		Regressor = LinearRegression()
		Regressor.fit(numpy.array(X_Norm).reshape(-1, 1),MovingAverages)
	#	print(str(X_Reg))
		y1= Regressor.predict(X_Reg[1])
		y2= Regressor.predict(X_Reg[2])
		TrendTest = (y2 -y1 )/(X_Reg[2] - X_Reg[1])
		if TrendTest > TrendTestLimit: print("Linear Regressor Trend: long")
		elif TrendTest < -TrendTestLimit: print("Linear Regressor Trend: Short")
		else: print("Linear Regressor Trend: Lateral/Undefined")      
        
# 2 - Análise de médias móveis
		#TrendShortMA = numpy.average(TrendAnalysisPrices[-TrendShortMAPeriod:])
		#TrendLongMA = numpy.average(TrendAnalysisPrices[-TrendLongMAPeriod:])  
        
		#print(str(MovingAverages))        
		#print("MovingAverages[-1]:" + str(MovingAverages[-1])+ "MovingAverages[(longTermengthOfMA-2):" + str(MovingAverages[(longTermengthOfMA-2)]) + "LTMovingAverage:" + str(LTMovingAverage) )
		if MovingAverages[-1] > MovingAverages[-2]: 
		 PolarisTradeLock = 'Long'
		 print("Moving Average Trend: long")

		if MovingAverages[-1] < MovingAverages[-2]: 
		 PolarisTradeLock = 'Short'
		 print("Moving Average Trend: Short")        

		if MovingAverages[-1] == MovingAverages[-2]: 
		 PolarisTradeLock = ''
		 print("Moving Average Trend: Undefined")        
		         		     
         
         
		if (PolarisTradeLock != UserTradeLock):
		  if ChangeTradeLock == True: 
		   TradeLock = PolarisTradeLock
		   print("Polaris is Switching TradeLock to "+ str(PolarisTradeLock))
		  elif ChangeTradeLock == False:  
		   TradeLock = ""
		   print("User defined trend does not match Polaris-defined Trend: Trading is halted")             

		if (PolarisTradeLock == UserTradeLock):  TradeLock = UserTradeLock       
          
		print("\n")  
		if (not startTime):
			time.sleep(int(period))


if __name__ == "__main__":
	main(sys.argv[1:])